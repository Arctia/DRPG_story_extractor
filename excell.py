
from deep_translator import GoogleTranslator
from dataloader import DataLoader
import openpyxl
import json
import main
import time
import os

class Translate():

	def __init__(self):
		self.translator = GoogleTranslator(source='auto', target='en')

	def translate(self, message):
		# add a translate flag
		try:
			result = self.translator.translate(message)
			return result
		except:
			print("\n[ERROR  ]: Translation error, waiting 30 secs...\n")
			time.sleep(30)
			return self.translate(message)

class Excell():

	name_jpc = 2
	name_enc = 5
	message_jpc = 3 # japan
	message_gtc = 4 # google
	message_mnc = 6 # manual

	offset = 2
	sz = 2
	step = 1

	row_pos = 2
	story_number = 1
	story_begin = 0

	def __init__(self, filename):
		self.translator = Translate()
		self.filename = filename
		self.file = openpyxl.load_workbook(filename)
		self.sheets = self.file.sheetnames

	def set_sheet(self, sheet_name):
		if not sheet_name in self.sheets:
			for s in self.sheets:
				switch = False
				for i in range(4):
					if sheet_name[i] != s[i]: 
						switch = True
						break
				if not switch:
					self.selected_sheet = self.file[s]
					self.selected_sheet.title = sheet_name
					self.reinit_sheet()
					print(f"[INFO    ]: Renaming {sheet_name} event")
					return False

			self.file.create_sheet(title=sheet_name)
			self.selected_sheet = self.file[sheet_name]# get_sheet_by_name(sheet_name)
			self.reinit_sheet()
			print(f"[INFO    ]: Add {sheet_name} event")
			return True
		else:
			return False

	def set_value(self, row:str, col:int, value:str):
		self.selected_sheet[row + str(col)].value = value

	def	set_value(self, row:int, col:int, value:str):
		self.selected_sheet.cell(row=row, column=col, value=value)

	def	get_value(self, row:str, col:int):
		value = self.selected_sheet[row + str(col)].value
		# exception caught
		return value

	def get_value(self, row:int, col:int):
		return self.selected_sheet.cell(row=row, column=col)

	def save(self):
		self.file.save(self.filename)

	def	reinit_sheet(self):
		self.row_pos = 3
		self.story_number = 1
		self.set_value(1, self.name_jpc, "Name (JP)")
		self.set_value(1, self.name_enc, "Name (Manual)")
		self.set_value(1, self.message_jpc, "JP")
		self.set_value(1, self.message_gtc, "Google translate")
		self.set_value(1, self.message_mnc, "Manual translation")

	def write_episode(self, episode_name):
		col = self.message_jpc
		row = self.row_pos

		self.set_value(row, 1, "Episode title:")
		self.set_value(row, col, episode_name)
		self.row_pos += 1

	def write_area(self, area_name):
		col = self.message_jpc
		row = self.row_pos

		print(f"	[INFO    ]: Add {area_name} area")
		self.set_value(row, 1, "Area name:")
		self.set_value(row, col, area_name)
		self.set_value(row, self.message_gtc, self.translate_sentence(area_name))
		self.row_pos += 2
		self.story_number = 1

	def write_story(self, story_name):
		col = self.message_jpc
		row = self.row_pos

		self.set_value(row, 1, f"Scene {self.story_number}:")
		self.set_value(row, col, story_name)
		self.set_value(row, self.message_gtc, self.translate_sentence(story_name))
		self.row_pos += 1
		self.story_begin = 0

		self.story_number += 1

	def	write_name(self, name):
		col = self.name_jpc
		row = self.row_pos

		if self.story_begin == 0:
			self.set_value(row, 1, "text:")
			self.story_begin = 1
		self.set_value(row, col, name)

	def write_message(self, message):
		col = self.message_jpc
		row = self.row_pos

		self.set_value(row, col, message)

	def write_translated_message(self, message):
		col = self.message_gtc
		row = self.row_pos

		self.set_value(row, col, self.translate_sentence(message))

		self.row_pos += self.step

	def translate_sentence(self, message):
		return self.translator.translate(message)

class DialogueExtractor(object):

	def __init__(self, file='japan.xlsx', jp=True):
		self.db = DataLoader(jp=jp)
		self.ex = Excell(file)
		self.start_cycle()

	def	get_name(self, t):
		options = ['chara1_name', 'chara2_name', 'chara3_name']
		for o in options:
			if t[o] != "":
				return t[o]

	def	story_talk_cycle(self, story):
		for talk in self.db.story_talk:
			if talk['m_story_id'] != story['id']: continue
			self.ex.write_name(self.get_name(talk))
			self.ex.write_message(talk['talk_text'].replace("\n", " "))
			self.ex.write_translated_message(talk['talk_text'].replace("\n", " "))

	def	story_cycle(self, area, ep):
		for story in self.db.story:
			divider = 100 if ep['event_type'] == 1 else 10
			if int(story['id'] / divider) != area['id']: continue

			self.ex.write_story(story['title'])
			self.story_talk_cycle(story)
			self.ex.row_pos += 1

	def	area_cycle(self, ep):
		area_num = 0
		for area in self.db.area:
			if area['m_episode_id'] != ep['m_episode_id']: continue
			area_num += 1
			area_name = area['name']
			self.ex.write_area(area['name'])
			self.story_cycle(area, ep)

	def start_cycle(self, tp='event'):
		event_types = [1, 15]
		story_type = self.db.event if tp == 'event' else self.db.episode 

		for ep in story_type:
			if not ep['event_type'] in event_types: continue
			sheet_name = f"{str(ep['id']).rjust(3, '0')}. {ep['resource_name']}"
			if not self.ex.set_sheet(sheet_name): continue
			self.area_cycle(ep)
			self.ex.save()

		self.ex.save()

if __name__ == '__main__':
	de = DialogueExtractor(file='japan.xlsx', jp=True)